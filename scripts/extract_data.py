#!/usr/bin/env python3
"""Extrai os dados da planilha para app/data.json (usado pelo web app).
Uso: python3 scripts/extract_data.py PLANEJAMENTO_SAFRA_2627_ORIGINAL.xlsx
"""
import openpyxl, json, sys, os

SRC = sys.argv[1] if len(sys.argv) > 1 else "PLANEJAMENTO_SAFRA_2627_ORIGINAL.xlsx"
OUT = os.path.join("app", "data.json")

wb = openpyxl.load_workbook(SRC, data_only=True)

def f(v):
    try: return round(float(v), 4)
    except: return 0.0
def s(v):
    return str(v).strip() if v is not None and str(v).strip() else ""

# PRODUTOS (PORTIFÓLIO): empresa, classe, produto, ativos, un, preço (S), estoque (T)
P = wb["PORTIFÓLIO"]
produtos = []
for r in range(4, 416):
    prod = s(P.cell(r, 3).value)
    if not prod: continue
    produtos.append(dict(empresa=s(P.cell(r,1).value), classe=s(P.cell(r,2).value),
        produto=prod, ativos=s(P.cell(r,4).value), un=s(P.cell(r,6).value),
        preco=f(P.cell(r,19).value), estoque=f(P.cell(r,20).value)))

# TALHÕES (ÁREA PLANTIO)
A = wb["ÁREA PLANTIO"]
talhoes = []
for r in range(2, A.max_row + 1):
    tid = s(A.cell(r,1).value)
    if not tid.upper().startswith("TL"): continue
    talhoes.append(dict(id=tid, nome=s(A.cell(r,2).value), empreendimento=s(A.cell(r,3).value),
        produtividade=f(A.cell(r,4).value), area=f(A.cell(r,5).value),
        emp_safrinha=s(A.cell(r,8).value), prod_safrinha=f(A.cell(r,9).value)))

# OPERAÇÕES por talhão (abas TLxx): dose em I, produto em C, classe em B, un em F
def extract_ops(ws, r0, r1):
    ops, cur = [], None
    for r in range(r0, r1 + 1):
        if s(ws.cell(r,1).value).upper().startswith("OPERA"):
            cur = {"nome": s(ws.cell(r,1).value), "itens": []}; ops.append(cur)
        prod = s(ws.cell(r,3).value)
        if prod and cur is not None:
            cur["itens"].append(dict(classe=s(ws.cell(r,2).value), produto=prod,
                dose=f(ws.cell(r,9).value), un=s(ws.cell(r,6).value)))
    return [o for o in ops if o["itens"]]

planos = {}
for t in talhoes:
    if t["id"] not in wb.sheetnames: continue
    ws = wb[t["id"]]
    planos[t["id"]] = dict(area=f(ws.cell(2,2).value), empreendimento=s(ws.cell(3,2).value),
        principal=extract_ops(ws, 10, 224), safrinha=extract_ops(ws, 229, 451))

# PREÇOS por cultura (DRE ORÇADA linha 7)
D = wb["DRE ORÇADA"]
precos_cultura = {}
for c in range(2, 15):
    emp, preco = s(D.cell(2,c).value), f(D.cell(7,c).value)
    if emp and preco: precos_cultura[emp] = round(preco, 2)

data = dict(safra="2026/2027", produtos=produtos, talhoes=talhoes,
            planos=planos, precos_cultura=precos_cultura)
os.makedirs("app", exist_ok=True)
json.dump(data, open(OUT, "w", encoding="utf-8"), ensure_ascii=False)
print(f"OK -> {OUT} | produtos={len(produtos)} talhões={len(talhoes)} planos={len(planos)}")
