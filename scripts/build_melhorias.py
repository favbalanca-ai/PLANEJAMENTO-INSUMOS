#!/usr/bin/env python3
"""Melhorias na planilha PLANEJAMENTO_SAFRA_2627.xlsx
1. Corrige o bug financeiro em PORTIFÓLIO!U (VOLUME INDICADO): R-T -> MAX(0;R-T)
2. Cria a aba 'DEMANDA DE COMPRAS' (lista acionavel; inclui MÁQUINAS)
3. Cria a aba 'COTAÇÃO FORNECEDOR' (itens a comprar agrupados por fornecedor,
   com colunas para o fornecedor preencher o preço cotado)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

SRC="PLANEJAMENTO_SAFRA_2627_ORIGINAL.xlsx"
OUT="PLANEJAMENTO_SAFRA_2627.xlsx"

wbv=openpyxl.load_workbook(SRC, data_only=True)
wb =openpyxl.load_workbook(SRC, data_only=False)
pv=wbv["PORTIFÓLIO"]; P=wb["PORTIFÓLIO"]

# ---- paleta ----
AZUL   ="FF16404D"   # banda titulo
AZUL2  ="FF2E7D8C"   # cabecalho tabela
TEAL   ="FF3E9AAA"   # banda fornecedor
CINZA  ="FFF4F6F8"   # zebra
VERDE  ="FFE3F3E0"
AMARELO="FFFFF2CC"
LINHA  ="FFB8C4CC"
WHITE  ="FFFFFFFF"
INK    ="FF33474F"
MUTED  ="FF64757D"
brl='R$ #,##0.00'; num='#,##0.00'
thin=Side(style="thin", color=LINHA)
border=Border(left=thin,right=thin,top=thin,bottom=thin)
def fnt(sz=10,b=False,c=INK,i=False): return Font(size=sz,bold=b,color=c,italic=i)
def fill(c): return PatternFill("solid", fgColor=c)

def fval(v):
    try: return float(v)
    except: return 0.0

# ---------------------------------------------------------------------------
# 1) CORRIGIR PORTIFÓLIO!U  ->  =MAX(0; R-T)
# ---------------------------------------------------------------------------
for r in range(4, 434):
    P.cell(r, 21).value = f"=MAX(0,R{r}-T{r})"

# ===========================================================================
# ABA 1 — DEMANDA DE COMPRAS  (inclui MÁQUINAS)
# ===========================================================================
sel=[]
for r in range(4, 416):
    prod=pv.cell(r,3).value
    if not (prod and str(prod).strip()): continue
    R=fval(pv.cell(r,18).value); T=fval(pv.cell(r,20).value)
    if R>0 or T>0: sel.append(r)
sel.sort(key=lambda r:(str(pv.cell(r,2).value or "").upper(),
                       str(pv.cell(r,3).value or "").upper()))

name="DEMANDA DE COMPRAS"
if name in wb.sheetnames: del wb[name]
ws=wb.create_sheet(name, 0)
ws.sheet_view.showGridLines=False

HDR=["CLASSE","FORNECEDOR","PRODUTO","UN","DEMANDA","ESTOQUE",
     "A COMPRAR","PREÇO UNIT.","VALOR TOTAL","STATUS"]
first=6; last=first+len(sel)-1

ws.merge_cells("A1:J1")
c=ws["A1"]; c.value="DEMANDA DE COMPRAS DE INSUMOS  —  SAFRA 2026/2027"
c.font=Font(color=WHITE,bold=True,size=16); c.fill=fill(AZUL)
c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
ws.row_dimensions[1].height=32
ws.merge_cells("A2:J2")
c=ws["A2"]; c.value=("A COMPRAR = MÁX(0; Demanda − Estoque).  Demanda e preço vêm da aba PORTIFÓLIO; "
                     "edite o ESTOQUE lá (coluna T) que esta lista atualiza sozinha.")
c.font=fnt(9,c=MUTED,i=True); c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
ws.row_dimensions[2].height=16

DATA=f"I{first}:I{last}"; GCOL=f"G{first}:G{last}"; HCOL=f"H{first}:H{last}"
def kpi(col,label,formula,fmt,vcolor):
    lc=ws.cell(3,col,label); lc.font=fnt(9,True,MUTED)
    lc.alignment=Alignment(horizontal="left",vertical="center")
    vc=ws.cell(4,col,formula); vc.font=Font(size=15,bold=True,color=vcolor)
    vc.alignment=Alignment(horizontal="left",vertical="center")
    if fmt: vc.number_format=fmt
kpi(1,"TOTAL A COMPRAR (R$)", f"=SUM({DATA})", brl, AZUL2)
kpi(3,"ITENS A COMPRAR",      f'=COUNTIF({GCOL},">0")', None, AZUL2)
kpi(5,"ITENS SEM PREÇO",      f'=SUMPRODUCT(({GCOL}>0)*({HCOL}<=0))', None, "FFB00020")
kpi(7,"VALOR EM ESTOQUE (R$)",f"=SUMPRODUCT(F{first}:F{last},H{first}:H{last})", brl, "FF2E7D32")
ws.row_dimensions[3].height=14; ws.row_dimensions[4].height=22

hrow=first-1
for i,h in enumerate(HDR,1):
    cell=ws.cell(hrow,i,h); cell.font=Font(color=WHITE,bold=True,size=10)
    cell.fill=fill(AZUL2); cell.border=border
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
ws.row_dimensions[hrow].height=28

for idx,r in enumerate(sel):
    row=first+idx
    ws.cell(row,1,f"='PORTIFÓLIO'!B{r}")
    ws.cell(row,2,f"='PORTIFÓLIO'!A{r}")
    ws.cell(row,3,f"='PORTIFÓLIO'!C{r}")
    ws.cell(row,4,f"='PORTIFÓLIO'!F{r}")
    ws.cell(row,5,f"='PORTIFÓLIO'!R{r}")
    ws.cell(row,6,f"='PORTIFÓLIO'!T{r}")
    ws.cell(row,7,f"=MAX(0,E{row}-F{row})")
    ws.cell(row,8,f"='PORTIFÓLIO'!S{r}")
    ws.cell(row,9,f"=G{row}*H{row}")
    ws.cell(row,10,f'=IF(G{row}>0,IF(H{row}<=0,"⚠ SEM PREÇO","COMPRAR"),'
                   f'IF(E{row}>0,"OK - ESTOQUE","SEM DEMANDA"))')
    for cc,fmt in ((5,num),(6,num),(7,num),(8,brl),(9,brl)):
        ws.cell(row,cc).number_format=fmt
    for col in range(1,11):
        cell=ws.cell(row,col); cell.border=border
        if col in (1,2,3,10): cell.alignment=Alignment(horizontal="left",vertical="center",indent=1)
        elif col==4: cell.alignment=Alignment(horizontal="center",vertical="center")
        else: cell.alignment=Alignment(horizontal="right",vertical="center")
    if idx%2==1:
        for col in range(1,11): ws.cell(row,col).fill=fill(CINZA)

tot=last+1
ws.cell(tot,3,"TOTAL GERAL").font=Font(bold=True,size=11,color=WHITE)
ws.cell(tot,7,f"=SUM(G{first}:G{last})").font=Font(bold=True,color=WHITE)
ws.cell(tot,9,f"=SUM(I{first}:I{last})").font=Font(bold=True,size=12,color=WHITE)
ws.cell(tot,9).number_format=brl; ws.cell(tot,7).number_format=num
for col in range(1,11):
    cell=ws.cell(tot,col); cell.fill=fill(AZUL); cell.border=border
    if cell.font is None or not cell.font.bold: cell.font=Font(color=WHITE)
ws.row_dimensions[tot].height=22

rng=f"A{first}:J{last}"
ws.conditional_formatting.add(rng,FormulaRule(formula=[f'$J{first}="COMPRAR"'],fill=fill(VERDE)))
ws.conditional_formatting.add(rng,FormulaRule(formula=[f'$J{first}="⚠ SEM PREÇO"'],fill=fill(AMARELO)))
ws.conditional_formatting.add(rng,FormulaRule(formula=[f'$J{first}="OK - ESTOQUE"'],fill=fill("FFEDEFF1")))

for col,w in {1:15,2:16,3:36,4:6,5:12,6:12,7:12,8:13,9:15,10:15}.items():
    ws.column_dimensions[get_column_letter(col)].width=w
ws.auto_filter.ref=f"A{hrow}:J{last}"
ws.freeze_panes=f"A{first}"

# resumo por classe (col L/M)
Lc=12
ws.cell(hrow,Lc,"RESUMO POR CLASSE").font=Font(color=WHITE,bold=True,size=10)
ws.merge_cells(start_row=hrow,start_column=Lc,end_row=hrow,end_column=Lc+1)
for cc in (Lc,Lc+1):
    ws.cell(hrow,cc).fill=fill(AZUL2)
ws.cell(hrow,Lc).alignment=Alignment(horizontal="center",vertical="center")
classes=[]
for r in sel:
    cl=str(pv.cell(r,2).value or "")
    if cl not in classes: classes.append(cl)
for i,cl in enumerate(classes):
    rr=hrow+1+i
    a=ws.cell(rr,Lc,cl); a.border=border; a.alignment=Alignment(horizontal="left",indent=1); a.font=fnt(10)
    b=ws.cell(rr,Lc+1,f'=SUMIF($A${first}:$A${last},L{rr},$I${first}:$I${last})')
    b.number_format=brl; b.border=border
rtot=hrow+1+len(classes)
ws.cell(rtot,Lc,"TOTAL").font=fnt(10,True)
ws.cell(rtot,Lc+1,f'=SUM(M{hrow+1}:M{rtot-1})').number_format=brl
ws.cell(rtot,Lc+1).font=fnt(10,True)
for cc in (Lc,Lc+1): ws.cell(rtot,cc).fill=fill(CINZA); ws.cell(rtot,cc).border=border
ws.column_dimensions[get_column_letter(Lc)].width=16
ws.column_dimensions[get_column_letter(Lc+1)].width=16

# ===========================================================================
# ABA 2 — COTAÇÃO FORNECEDOR  (itens a comprar, agrupados por fornecedor)
# ===========================================================================
cot=[]  # insumos (sem maquinas) com A COMPRAR > 0
for r in range(4,416):
    prod=pv.cell(r,3).value
    if not (prod and str(prod).strip()): continue
    if str(pv.cell(r,2).value or "").strip().upper().startswith("MÁQUINA"): continue
    R=fval(pv.cell(r,18).value); T=fval(pv.cell(r,20).value)
    if max(0,R-T)>0: cot.append(r)

# agrupa por fornecedor
groups={}
for r in cot:
    forn=pv.cell(r,1).value
    forn=str(forn).strip() if forn and str(forn).strip() else "(SEM FORNECEDOR)"
    groups.setdefault(forn,[]).append(r)
order=sorted(groups, key=lambda k:(k=="(SEM FORNECEDOR)", k))  # sem fornecedor por ultimo

name2="COTAÇÃO FORNECEDOR"
if name2 in wb.sheetnames: del wb[name2]
cs=wb.create_sheet(name2, 1)
cs.sheet_view.showGridLines=False

CH=["PRODUTO","CLASSE","UN","QTD A COMPRAR","PREÇO REF. (R$)","VALOR REF. (R$)",
    "PREÇO COTADO (R$)","VALOR COTADO (R$)"]
NC=len(CH)
lastcol=get_column_letter(NC)

cs.merge_cells(f"A1:{lastcol}1")
c=cs["A1"]; c.value="COTAÇÃO DE INSUMOS POR FORNECEDOR  —  SAFRA 2026/2027"
c.font=Font(color=WHITE,bold=True,size=16); c.fill=fill(AZUL)
c.alignment=Alignment(horizontal="left",vertical="center",indent=1); cs.row_dimensions[1].height=32
cs.merge_cells(f"A2:{lastcol}2")
c=cs["A2"]; c.value=("Quantidades já descontam o estoque. Preencha 'PREÇO COTADO' para comparar "
                     "com o preço de referência; o 'VALOR COTADO' calcula sozinho.")
c.font=fnt(9,c=MUTED,i=True); c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
cs.row_dimensions[2].height=16

rowptr=4
subtotal_ref_cells=[]; subtotal_cot_cells=[]
for forn in order:
    rows=sorted(groups[forn], key=lambda r:str(pv.cell(r,3).value or "").upper())
    # banda fornecedor
    cs.merge_cells(start_row=rowptr,start_column=1,end_row=rowptr,end_column=NC)
    fc=cs.cell(rowptr,1,f"FORNECEDOR:  {forn}    ({len(rows)} itens)")
    fc.font=Font(color=WHITE,bold=True,size=11); fc.fill=fill(TEAL)
    fc.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    cs.row_dimensions[rowptr].height=22
    rowptr+=1
    # cabecalho
    for i,h in enumerate(CH,1):
        cell=cs.cell(rowptr,i,h); cell.font=Font(color=WHITE,bold=True,size=9)
        cell.fill=fill(AZUL2); cell.border=border
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    cs.row_dimensions[rowptr].height=26
    rowptr+=1
    gstart=rowptr
    for j,r in enumerate(rows):
        cs.cell(rowptr,1,f"='PORTIFÓLIO'!C{r}")
        cs.cell(rowptr,2,f"='PORTIFÓLIO'!B{r}")
        cs.cell(rowptr,3,f"='PORTIFÓLIO'!F{r}")
        cs.cell(rowptr,4,f"='PORTIFÓLIO'!U{r}")          # qtd = MAX(0,R-T)
        cs.cell(rowptr,5,f"='PORTIFÓLIO'!S{r}")          # preço ref
        cs.cell(rowptr,6,f"=D{rowptr}*E{rowptr}")        # valor ref
        cs.cell(rowptr,7,None)                            # preço cotado (fornecedor preenche)
        cs.cell(rowptr,8,f"=D{rowptr}*G{rowptr}")        # valor cotado
        cs.cell(rowptr,4).number_format=num
        for cc in (5,6,7,8): cs.cell(rowptr,cc).number_format=brl
        for col in range(1,NC+1):
            cell=cs.cell(rowptr,col); cell.border=border
            if col in (1,2): cell.alignment=Alignment(horizontal="left",vertical="center",indent=1)
            elif col==3: cell.alignment=Alignment(horizontal="center",vertical="center")
            else: cell.alignment=Alignment(horizontal="right",vertical="center")
        if col==7: pass
        cs.cell(rowptr,7).fill=fill(AMARELO)             # campo a preencher destacado
        if j%2==1:
            for col in (1,2,3,4,5,6,8): cs.cell(rowptr,col).fill=fill(CINZA)
        rowptr+=1
    # subtotal fornecedor
    cs.cell(rowptr,1,f"Subtotal {forn}").font=fnt(10,True)
    cs.cell(rowptr,6,f"=SUM(F{gstart}:F{rowptr-1})").number_format=brl
    cs.cell(rowptr,8,f"=SUM(H{gstart}:H{rowptr-1})").number_format=brl
    for col in range(1,NC+1):
        cell=cs.cell(rowptr,col); cell.fill=fill("FFE8EDF0"); cell.border=border
        if cell.font is None or not cell.font.bold: cell.font=fnt(10,True)
    subtotal_ref_cells.append(f"F{rowptr}"); subtotal_cot_cells.append(f"H{rowptr}")
    cs.row_dimensions[rowptr].height=20
    rowptr+=2  # espacador

# total geral
cs.cell(rowptr,1,"TOTAL GERAL").font=Font(bold=True,size=12,color=WHITE)
cs.cell(rowptr,6,"="+"+".join(subtotal_ref_cells)).number_format=brl
cs.cell(rowptr,8,"="+"+".join(subtotal_cot_cells)).number_format=brl
for col in range(1,NC+1):
    cell=cs.cell(rowptr,col); cell.fill=fill(AZUL); cell.border=border
    if cell.font is None or not cell.font.bold: cell.font=Font(color=WHITE,bold=True)
cs.row_dimensions[rowptr].height=24

for col,w in {1:36,2:15,3:6,4:14,5:15,6:16,7:16,8:16}.items():
    cs.column_dimensions[get_column_letter(col)].width=w
cs.freeze_panes="A4"

# ---- recalculo ao abrir ----
try: wb.calculation.fullCalcOnLoad=True
except Exception: pass

wb.save(OUT)
print(f"OK -> {OUT}")
print(f"DEMANDA: {len(sel)} itens (com máquinas) | linhas {first}-{last}")
print(f"COTAÇÃO: {len(cot)} itens em {len(order)} fornecedores")
